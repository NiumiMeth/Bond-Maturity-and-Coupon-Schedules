import streamlit as st
st.set_page_config(layout="wide", page_title="Coupon Schedules", page_icon="🏛️")

import pandas as pd
import numpy as np
import io
import re
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
  font-family: 'DM Sans', sans-serif;
  background: #F5F1EB !important;
  color: #1C2B1E !important;
}
.stApp {
  background: linear-gradient(160deg, #F5F1EB 0%, #EDE8DF 60%, #E8E2D6 100%) !important;
}
/* Override Streamlit defaults */
.st-* { background-color: transparent !important; }
[data-testid="stForm"] { background: #FFFFFF !important; border: 1px solid #D4CCC0 !important; }

#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 1440px; }

/* ── PAGE HEADER ── */
.page-header {
  border-bottom: 2px solid #2A4A30;
  padding-bottom: 1.5rem;
  margin-bottom: 2.5rem;
}
.page-header .sub {
  font-size: .72rem;
  color: #7A8C7E;
  letter-spacing: .22em;
  text-transform: uppercase;
  margin-bottom: .4rem;
}
.page-header h1 {
  font-family: 'Playfair Display', serif;
  font-size: 2.5rem;
  font-weight: 700;
  color: #1C2B1E;
  margin: 0;
  line-height: 1.15;
}
.page-header .ts {
  font-family: 'DM Mono', monospace;
  font-size: .72rem;
  color: #7A8C7E;
  margin-top: .5rem;
}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
  background: #F9F6F0 !important;
  border-radius: 6px 6px 0 0;
  border: 1px solid #C8BFB0;
  border-bottom: none;
  gap: 0;
}
.stTabs [data-baseweb="tab"] {
  font-family: 'DM Mono', monospace !important;
  font-size: .73rem !important;
  letter-spacing: .12em !important;
  text-transform: uppercase !important;
  color: #7A8C7E !important;
  padding: .75rem 1.6rem !important;
  background: #F9F6F0 !important;
}
.stTabs [aria-selected="true"] {
  color: #1C2B1E !important;
  border-bottom: 2px solid #2A4A30 !important;
  background: #FFFFFF !important;
}
.stTabs [data-baseweb="tab-panel"] {
  background: #FFFFFF !important;
  border: 1px solid #C8BFB0;
  border-top: none;
  border-radius: 0 0 6px 6px;
  padding: 1.5rem;
}

/* ── FILE UPLOADER ── */
[data-testid="stFileUploader"] {
  background: #FFFFFF !important;
  border: 1.5px dashed #B0A898;
  border-radius: 6px;
  padding: .5rem;
}
[data-testid="stFileUploader"]:hover { border-color: #2A4A30; }
[data-testid="stFileUploaderDropzone"] { background: #F9F6F0 !important; }
[data-testid="stFileUploaderDropzone"] p { color: #7A8C7E; }

/* ── METRICS ── */
.metric-row { display: flex; gap: 1.2rem; margin-bottom: 2.5rem; flex-wrap: wrap; }
.mc {
  flex: 1; min-width: 160px;
  background: #FFFFFF;
  border: 1px solid #D4CCBf;
  border-top: 3px solid #2A4A30;
  border-radius: 6px;
  padding: 1.2rem 1.5rem;
  box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.mc.warn { border-top-color: #C47B1A; }
.mc.danger { border-top-color: #C43B2A; }
.mc .lbl {
  font-size: .68rem;
  font-weight: 600;
  letter-spacing: .18em;
  text-transform: uppercase;
  color: #7A8C7E;
  margin-bottom: .5rem;
}
.mc .val {
  font-family: 'DM Mono', monospace;
  font-size: 1.5rem;
  font-weight: 500;
  color: #1C2B1E;
  line-height: 1;
}
.mc.warn .val { color: #C47B1A; }
.mc.danger .val { color: #C43B2A; }
.mc .sub { font-size: .72rem; color: #A09890; margin-top: .4rem; }

/* ── SECTION TITLE ── */
.sec-title {
  font-family: 'Playfair Display', serif;
  font-size: 1.15rem;
  color: #1C2B1E;
  letter-spacing: .02em;
  margin-bottom: 1rem;
  padding-bottom: .5rem;
  border-bottom: 1px solid #C8BFB0;
}

/* ── BADGE ── */
.badge {
  display: inline-block;
  font-family: 'DM Mono', monospace;
  font-size: .66rem;
  padding: .18rem .6rem;
  border-radius: 3px;
  font-weight: 500;
  letter-spacing: .06em;
}
.bg  { background: #E8F0E8; color: #2A4A30; border: 1px solid #B8D4B8; }
.bw  { background: #FEF3E2; color: #C47B1A; border: 1px solid #F5D89A; }
.br  { background: #FEEBEA; color: #C43B2A; border: 1px solid #F5B8B4; }
.bb  { background: #E8F0F8; color: #2A5090; border: 1px solid #B4CDE8; }

/* ── TABLE ── */
.tbl-wrap {
  border-radius: 6px;
  overflow: hidden;
  border: 1px solid #D4CCC0;
  margin-bottom: .5rem;
  box-shadow: 0 1px 3px rgba(0,0,0,.05);
}
.tbl { width: 100%; border-collapse: collapse; font-size: .84rem; }
.tbl thead th {
  background: #2A4A30;
  color: #D4E8D4;
  font-size: .66rem;
  font-weight: 600;
  letter-spacing: .16em;
  text-transform: uppercase;
  padding: .7rem 1rem;
  text-align: left;
}
.tbl thead th.num { text-align: right; }
.tbl tbody tr { border-bottom: 1px solid #EDE8DF; transition: background .12s; }
.tbl tbody tr:last-child { border-bottom: none; }
.tbl tbody tr:hover { background: #F0EBE0; }
.tbl tbody td { padding: .65rem 1rem; color: #1C2B1E; }
.tbl tbody td.mono {
  font-family: 'DM Mono', monospace;
  font-size: .8rem;
  color: #2A5090;
}
.tbl tbody td.dt {
  font-family: 'DM Mono', monospace;
  font-size: .8rem;
  color: #5A7A5E;
}
.tbl tbody td.amt {
  font-family: 'DM Mono', monospace;
  text-align: right;
  color: #1C2B1E;
}
.tbl tbody td.warn {
  color: #C47B1A;
  font-family: 'DM Mono', monospace;
}
.tbl tfoot tr {
  background: #EDE8DF;
  border-top: 1.5px solid #2A4A30;
}
.tbl tfoot td {
  padding: .7rem 1rem;
  font-family: 'DM Mono', monospace;
  font-size: .82rem;
  font-weight: 600;
  color: #1C2B1E;
  letter-spacing: .06em;
}
.tbl tfoot td.ta { text-align: right; font-size: .9rem; }

/* ── EXPANDER ── */
[data-testid="stExpander"] {
  background: #FFFFFF !important;
  border: 1px solid #D4CCC0 !important;
  border-radius: 6px !important;
  margin-bottom: .6rem !important;
  box-shadow: 0 1px 3px rgba(0,0,0,.04) !important;
}
[data-testid="stExpander"] summary {
  font-family: 'DM Sans', sans-serif !important;
  font-weight: 600 !important;
  font-size: .88rem !important;
  color: #1C2B1E !important;
  letter-spacing: .04em !important;
  text-transform: uppercase !important;
  padding: .9rem 1.2rem !important;
  background: #F9F6F0 !important;
}
[data-testid="stExpander"] summary:hover { 
  color: #2A4A30 !important; 
  background: #EDE8DF !important;
}
[data-testid="stExpander"] [role="region"] {
  background: #FFFFFF !important;
}

/* ── BUTTONS ── */
.stDownloadButton button {
  background: #2A4A30 !important;
  border: none !important;
  color: #F5F1EB !important;
  font-family: 'DM Mono', monospace !important;
  font-size: .76rem !important;
  letter-spacing: .12em !important;
  text-transform: uppercase !important;
  border-radius: 4px !important;
  padding: .55rem 1.4rem !important;
}
.stDownloadButton button:hover { background: #1C3320 !important; }

/* Override any Streamlit dark defaults for text inputs and select boxes */
input, select, textarea {
  background: #FFFFFF !important;
  color: #1C2B1E !important;
  border: 1px solid #D4CCC0 !important;
}
input:focus, select:focus, textarea:focus {
  background: #FFFFFF !important;
  color: #1C2B1E !important;
  border-color: #2A4A30 !important;
}

/* ── INFO / WARN BOXES ── */
.info-box {
  background: #FFFFFF;
  border: 1.5px dashed #C8BFB0;
  border-radius: 8px;
  padding: 3rem 2rem;
  text-align: center;
  color: #A09890;
}
.info-box .ico { font-size: 2.5rem; margin-bottom: 1rem; }
.info-box p { margin: 0; font-size: .9rem; letter-spacing: .04em; }

.warn-box {
  background: #FEF8EE;
  border: 1px solid #F5D89A;
  border-left: 3px solid #C47B1A;
  border-radius: 5px;
  padding: .85rem 1.2rem;
  font-size: .83rem;
  color: #7A4E10;
  margin: .8rem 0;
}
.step-banner {
  background: #FEF8EE;
  border: 1px solid #F5D89A;
  border-left: 3px solid #C47B1A;
  border-radius: 5px;
  padding: 1rem 1.2rem;
  margin: 1rem 0;
  font-size: .83rem;
  color: #7A4E10;
}
.step-banner strong { color: #C47B1A; }

.grp-lbl {
  font-size: .68rem;
  letter-spacing: .18em;
  text-transform: uppercase;
  color: #7A8C7E;
  margin: 1rem 0 .5rem 0;
  display: flex;
  align-items: center;
  gap: .5rem;
}
.grp-lbl::after { content: ''; flex: 1; height: 1px; background: #D4CCC0; }

/* view tabs within dashboard */
.view-tab-row { display:flex; gap:.5rem; margin-bottom:1.5rem; }
.view-tab-btn {
  font-family:'DM Mono',monospace; font-size:.73rem; letter-spacing:.1em;
  text-transform:uppercase; padding:.5rem 1.2rem; border-radius:4px; cursor:pointer;
  border:1px solid #C8BFB0; background:#FFFFFF; color:#7A8C7E; transition:all .15s;
}
.view-tab-btn.active {
  background:#2A4A30; color:#F5F1EB; border-color:#2A4A30;
}
</style>
""", unsafe_allow_html=True)

# ─── CALCULATION LOGIC ────────────────────────────────────────────────────────

REF_DATE = date(2026, 3, 31)
APR_BASE = date(2026, 4, 1)
MAY_BASE = date(2026, 5, 1)


def parse_series(series: str):
    series = str(series).strip()
    pcts = re.findall(r'(\d{1,2}\.\d{2})%', series)
    if not pcts:
        pcts = re.findall(r'(\d{1,2}(?:\.\d+)?)%', series)
    rates = [float(p) / 100 for p in pcts]
    if len(rates) == 0:
        return 'normal', [0.0]
    elif len(rates) == 1:
        return 'normal', rates
    elif len(rates) == 2:
        return 'step2', rates
    else:
        return 'step3', rates


def calc_remaining_coupons(maturity: date) -> float:
    base = APR_BASE if maturity.day >= 15 else MAY_BASE
    months = (maturity.year - base.year) * 12 + (maturity.month - base.month)
    return round(months / 6, 2)


def process_row(row: dict) -> dict:
    mat_raw = row.get('Maturity Date')
    mat = None
    if isinstance(mat_raw, str):
        for fmt in ['%d-%b-%y', '%d-%b-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                mat = datetime.strptime(mat_raw.strip(), fmt).date()
                break
            except:
                pass
    elif isinstance(mat_raw, datetime):
        mat = mat_raw.date()
    elif isinstance(mat_raw, date):
        mat = mat_raw
    elif pd.notna(mat_raw):
        try:
            mat = pd.to_datetime(mat_raw).date()
        except:
            mat = None

    series = str(row.get('Series', ''))
    face   = float(str(row.get('Face Value (Rs Mn)', 0)).replace(',', '')) if pd.notna(row.get('Face Value (Rs Mn)')) else 0.0

    bond_type, rates = parse_series(series)
    current_rate     = rates[-1] if rates else 0.0

    remaining   = calc_remaining_coupons(mat) if mat else None
    semi_coupon = round(face * current_rate / 2, 5) if face and current_rate else None
    total_pay   = round(semi_coupon * remaining, 4) if semi_coupon and remaining else None

    if bond_type == 'normal':
        coupon_rate_str = f"{current_rate*100:.2f}%"
    else:
        coupon_rate_str = '/'.join(f"{r*100:.2f}%" for r in rates)

    return {
        'Maturity Date':              mat.strftime('%d-%b-%y') if mat else str(mat_raw),
        'ISIN':                       row.get('ISIN', ''),
        'Series':                     series,
        'Face Value (Rs Mn)':         face,
        'Coupon Rate':                coupon_rate_str,
        'Current Rate':               current_rate,
        'Remaining Coupons':          remaining,
        'Semiannual Coupon Payment':  semi_coupon,
        'Bond Type':                  bond_type,
        'All Rates':                  rates,
        '_mat':                       mat,
    }


def parse_uploaded_file(uploaded_file):
    name = uploaded_file.name.lower()
    raw_bytes = uploaded_file.read()

    REQUIRED = ['maturity date', 'isin', 'series', 'face value']

    def find_header_row(df_raw):
        for i, row in df_raw.iterrows():
            vals = [str(v).lower().strip() for v in row.values]
            matches = sum(any(req in v for v in vals) for req in REQUIRED)
            if matches >= 3:
                return i
        return None

    if name.endswith('.csv'):
        df_raw = pd.read_csv(io.BytesIO(raw_bytes), header=None)
        hrow   = find_header_row(df_raw)
        df     = pd.read_csv(io.BytesIO(raw_bytes), skiprows=hrow) if hrow is not None else pd.read_csv(io.BytesIO(raw_bytes))

    elif name.endswith(('.xlsx', '.xls')):
        df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=None)
        hrow   = find_header_row(df_raw)
        df     = pd.read_excel(io.BytesIO(raw_bytes), skiprows=hrow) if hrow is not None else pd.read_excel(io.BytesIO(raw_bytes))
    else:
        return None, "unsupported"

    col_map = {}
    mapped_targets = set()
    for col in df.columns:
        cl = str(col).lower().strip()
        target = None
        if 'maturity' in cl or ('date' in cl and 'mat' in cl):
            target = 'Maturity Date'
        elif cl == 'isin' or 'isin' in cl:
            target = 'ISIN'
        elif 'series' in cl:
            target = 'Series'
        elif 'face' in cl or ('value' in cl and 'face' in cl):
            target = 'Face Value (Rs Mn)'

        if target and target not in mapped_targets:
            col_map[col] = target
            mapped_targets.add(target)

    df.rename(columns=col_map, inplace=True)

    if 'Maturity Date' in df.columns:
        # Fill missing maturity dates with the previous row's date
        df['Maturity Date'] = df['Maturity Date'].ffill()
        
        # Remove rows where maturity date is still empty (first rows with no prior date)
        df = df[df['Maturity Date'].notna()]
        df = df[df['Maturity Date'].astype(str).str.strip() != '']
        df = df[df['Maturity Date'].astype(str).str.strip().str.lower() != 'maturity date']

    return df, "ok"


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def build_excel(results: list) -> bytes:
    wb = Workbook()

    HDR_BG   = "2A4A30"
    HDR_FG   = "D4E8D4"
    ALT1_BG  = "FFFFFF"
    ALT2_BG  = "F5F1EB"
    WARN_BG  = "FEF8EE"
    WARN_FG  = "7A4E10"
    ERR_BG   = "FEEBEA"
    ERR_FG   = "7A2010"
    TOT_BG   = "EDE8DF"
    TOT_FG   = "1C2B1E"
    HDR_TXT  = "1C2B1E"

    def s(style='thin', color="C8BFB0"):
        return Side(style=style, color=color)
    def tb():
        return Border(left=s(), right=s(), top=s(), bottom=s())
    def mb():
        return Border(left=s('medium','2A4A30'), right=s('medium','2A4A30'),
                      top=s('medium','2A4A30'), bottom=s('medium','2A4A30'))

    normal = [r for r in results if r['Bond Type'] == 'normal']
    step2  = [r for r in results if r['Bond Type'] == 'step2']
    step3  = [r for r in results if r['Bond Type'] == 'step3']

    sheet_data = [("All Bonds", results), ("Normal Bonds", normal)]
    if step2: sheet_data.append(("2-Step Bonds", step2))
    if step3: sheet_data.append(("3-Step Bonds", step3))

    first = True
    for sheet_name, data in sheet_data:
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        ws.merge_cells('A1:I1')
        c = ws['A1']
        c.value = f"Treasury Bonds Outstanding — As at 31 March 2026 ({sheet_name})"
        c.font = Font(name='Arial', bold=True, size=12, color=HDR_FG)
        c.fill = PatternFill('solid', fgColor=HDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        if sheet_name != "Normal Bonds" and (step2 or step3):
            ws.merge_cells('A2:I2')
            c = ws['A2']
            c.value = "⚠  Stepped coupon bonds present — calculations use current/final rate only. Verify step dates."
            c.font = Font(name='Arial', size=9, color=WARN_FG)
            c.fill = PatternFill('solid', fgColor=WARN_BG)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 18
            hdr_row = 3
        else:
            hdr_row = 2

        headers = ["Maturity Date","ISIN","Series","Face Value (Rs Mn)",
                   "Coupon Rate","Bond Type","Remaining Coupons","Semiannual Coupon Payment"]

        ws.row_dimensions[hdr_row].height = 32
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font = Font(name='Arial', bold=True, size=9, color=HDR_FG)
            c.fill = PatternFill('solid', fgColor=HDR_BG)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = tb()

        for ri, row in enumerate(data):
            er = ri + hdr_row + 1
            bt = row['Bond Type']
            if bt == 'step2':
                bg, fc = WARN_BG, WARN_FG
            elif bt == 'step3':
                bg, fc = ERR_BG, ERR_FG
            else:
                bg, fc = (ALT1_BG if ri % 2 == 0 else ALT2_BG), HDR_TXT
            fill = PatternFill('solid', fgColor=bg)

            vals = [
                row['Maturity Date'], row['ISIN'], row['Series'],
                row['Face Value (Rs Mn)'], row['Coupon Rate'],
                {'normal':'Normal','step2':'2-Step ⚠','step3':'3-Step ⚠⚠'}.get(bt, bt),
                row['Remaining Coupons'],
                row['Semiannual Coupon Payment'],
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=er, column=ci, value=val)
                c.fill = fill; c.border = tb()
                c.font = Font(name='Arial', size=9, color=fc)
                if ci == 4:
                    c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right')
                elif ci == 5:
                    c.alignment = Alignment(horizontal='center')
                elif ci == 7:
                    c.number_format = '0.00'; c.alignment = Alignment(horizontal='center')
                elif ci == 8:
                    c.number_format = '#,##0.000'; c.alignment = Alignment(horizontal='right')
                else:
                    c.alignment = Alignment(horizontal='center' if ci == 1 else 'left')

        tr = hdr_row + len(data) + 1
        ws.merge_cells(f'A{tr}:C{tr}')
        c = ws.cell(row=tr, column=1, value="TOTAL")
        c.font = Font(name='Arial', bold=True, size=9, color=TOT_FG)
        c.fill = PatternFill('solid', fgColor=TOT_BG)
        c.alignment = Alignment(horizontal='center')
        c.border = mb()
        for ci, col_idx in [(4,4),(8,8)]:
            cl = get_column_letter(ci)
            c = ws.cell(row=tr, column=ci)
            c.value = f'=SUM({cl}{hdr_row+1}:{cl}{tr-1})'
            c.font = Font(name='Arial', bold=True, size=9, color=TOT_FG)
            c.fill = PatternFill('solid', fgColor=TOT_BG)
            c.border = mb()
            c.alignment = Alignment(horizontal='right')
            c.number_format = '#,##0.00' if ci == 4 else '#,##0.000'
        for ci in (5,6,7):
            c = ws.cell(row=tr, column=ci)
            c.value = ''
            c.fill = PatternFill('solid', fgColor=TOT_BG)
            c.border = mb()
        for ci, w in enumerate([13,18,24,16,20,12,14,22,24], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = f'A{hdr_row+1}'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── TABLE RENDER ─────────────────────────────────────────────────────────────

def make_table(rows, show_step_col=False):
    cols = ["ISIN","Series","Maturity Date","Face Value (Rs Mn)",
            "Coupon Rate","Remaining Coupons","Semiannual Coupon Payment"]
    if show_step_col:
        cols = ["Bond Type"] + cols

    hdr = "".join(
        f'<th class="num">{c}</th>' if c in ("Face Value (Rs Mn)","Semiannual Coupon Payment","Remaining Coupons")
        else f'<th>{c}</th>' for c in cols
    )

    body = ""
    for r in rows:
        bt = r['Bond Type']
        body += "<tr>"
        for c in cols:
            v = r.get(c, '')
            if c == "Bond Type":
                badge = {'normal':'<span class="badge bg">Normal</span>',
                         'step2':'<span class="badge bw">2-Step ⚠</span>',
                         'step3':'<span class="badge br">3-Step ⚠⚠</span>'}.get(bt, bt)
                body += f'<td>{badge}</td>'
            elif c == "ISIN":
                body += f'<td class="mono">{v}</td>'
            elif c == "Maturity Date":
                body += f'<td class="dt">{v}</td>'
            elif c in ("Face Value (Rs Mn)","Semiannual Coupon Payment"):
                try:    body += f'<td class="amt">{float(v):,.3f}</td>'
                except: body += f'<td class="amt">{v}</td>'
            elif c == "Remaining Coupons":
                try:    body += f'<td class="amt">{float(v):.2f}</td>'
                except: body += f'<td class="amt">{v}</td>'
            elif c == "Coupon Rate" and bt != 'normal':
                body += f'<td class="warn">{v}</td>'
            else:
                body += f'<td>{v}</td>'
        body += "</tr>"

    n = len(rows)
    ncols = len(cols)
    total_semi = sum(r.get('Semiannual Coupon Payment') or 0 for r in rows)
    foot = f'<td colspan="{ncols-1}">Total · {n} bond{"s" if n!=1 else ""}</td><td class="ta">{total_semi:,.3f}</td>'

    current_year_bonds = [r for r in rows if r.get('_mat') and r['_mat'].year == 2026]
    foot2 = ""
    if current_year_bonds:
        cy_principal = sum(r.get('Face Value (Rs Mn)') or 0 for r in current_year_bonds)
        cy_total = total_semi + cy_principal
        foot2 = f'<tr><td colspan="{ncols-1}">Maturing in 2026 (Coupon + Principal)</td><td class="ta">{cy_total:,.3f}</td></tr>'

    return f"""
<div class="tbl-wrap">
<table class="tbl">
<thead><tr>{hdr}</tr></thead>
<tbody>{body}</tbody>
<tfoot><tr>{foot}</tr>{foot2}</tfoot>
</table>
</div>"""


# ─── NORMAL BONDS VIEW (no step-down ISINs) ───────────────────────────────────

def render_normal_view(results: list):
    normal = [r for r in results if r['Bond Type'] == 'normal']
    n_stepped = len([r for r in results if r['Bond Type'] != 'normal'])

    st.markdown('<div class="sec-title">Normal Bonds — Schedule by Maturity Month</div>', unsafe_allow_html=True)

    if normal:
        for r in normal:
            if r['_mat']:
                r['_month']    = r['_mat'].strftime('%B')
                r['_monthnum'] = r['_mat'].month
            else:
                r['_month']    = 'Unknown'
                r['_monthnum'] = 99

        # Filter to only show bonds with valid maturity dates
        normal_with_dates = [r for r in normal if r['_monthnum'] != 99]
        
        if not normal_with_dates:
            st.markdown('<div class="info-box"><div class="ico">📋</div><p>No normal bonds found.</p></div>', unsafe_allow_html=True)
            return
        
        months = sorted(set((r['_monthnum'], r['_month']) for r in normal_with_dates), key=lambda x: x[0])

        for mnum, mname in months:
            month_rows = [r for r in normal_with_dates if r['_monthnum'] == mnum]
            with st.expander(f"{mname}  ·  {len(month_rows)} bond{'s' if len(month_rows)!=1 else ''}"):
                first_rows = [r for r in month_rows if r['_mat'] and r['_mat'].day == 1]
                fif_rows   = [r for r in month_rows if r['_mat'] and r['_mat'].day == 15]
                other_rows = [r for r in month_rows if r['_mat'] and r['_mat'].day not in (1,15)]

                if first_rows:
                    st.markdown('<div class="grp-lbl"><span class="badge bg">1st</span></div>', unsafe_allow_html=True)
                    st.markdown(make_table(first_rows), unsafe_allow_html=True)
                if fif_rows:
                    st.markdown('<div class="grp-lbl"><span class="badge bb">15th</span></div>', unsafe_allow_html=True)
                    st.markdown(make_table(fif_rows), unsafe_allow_html=True)
                if other_rows:
                    st.markdown('<div class="grp-lbl"><span class="badge bw">Other Date</span></div>', unsafe_allow_html=True)
                    st.markdown(make_table(other_rows), unsafe_allow_html=True)
    else:
        st.markdown('<div class="info-box"><div class="ico">📋</div><p>No normal bonds found.</p></div>', unsafe_allow_html=True)


# ─── ALL BONDS VIEW (including step-down ISINs) ───────────────────────────────

def render_all_view(results: list):
    step2   = [r for r in results if r['Bond Type'] == 'step2']
    step3   = [r for r in results if r['Bond Type'] == 'step3']
    stepped = step2 + step3

    if not stepped:
        st.markdown('<div class="info-box"><div class="ico">✅</div><p>No stepped-coupon bonds found in this file.</p></div>', unsafe_allow_html=True)
        return

    st.markdown(f"""
    <div class="step-banner">
      ⚠ &nbsp; <strong>{len(stepped)}</strong> stepped-coupon bond{'s' if len(stepped)>1 else ''} detected.
      Calculations use the <strong>current/final rate only</strong> — step dates required for full accuracy.
    </div>
    """, unsafe_allow_html=True)

    if step2:
        st.markdown('<div class="sec-title" style="color:#C47B1A;border-bottom-color:#F5D89A">2-Step Bonds ⚠</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="warn-box">
          These bonds have <strong>two coupon rates</strong> (e.g. 12.00%→09.00%).
          Calculations use the <strong>current/final rate</strong> only.
        </div>
        """, unsafe_allow_html=True)
        st.markdown(make_table(step2, show_step_col=True), unsafe_allow_html=True)

    if step3:
        st.markdown('<div class="sec-title" style="color:#C43B2A;border-bottom-color:#F5B8B4">3-Step Bonds ⚠⚠</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="warn-box" style="border-left-color:#C43B2A;color:#7A2010;background:#FEEBEA;border-color:#F5B8B4">
          These bonds have <strong>three coupon rates</strong> and require both step dates for accurate calculation.
          Figures shown use the <strong>final/lowest rate</strong> only.
        </div>
        """, unsafe_allow_html=True)
        st.markdown(make_table(step3, show_step_col=True), unsafe_allow_html=True)


# ─── DASHBOARD WRAPPER ────────────────────────────────────────────────────────

def render_dashboard(results: list, key_suffix: str = "xl"):
    # Inner view tabs
    view_tab_normal, view_tab_all = st.tabs([
        "📅  View Bonds",
        "⚠  Step-Down ISINs",
    ])

    with view_tab_normal:
        render_normal_view(results)

    with view_tab_all:
        render_all_view(results)

    st.markdown("<br>", unsafe_allow_html=True)
    xlsx = build_excel(results)
    st.download_button(
        label="⬇  Download Complete Excel",
        data=xlsx,
        file_name="T_Bonds_Calculated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_excel_{key_suffix}"
    )


# ─── MAIN UI ──────────────────────────────────────────────────────────────────

st.markdown("""
<div class="page-header">
  <div class="sub">Central Bank of Sri Lanka · Fixed Income Operations</div>
  <h1>Bond Maturity and Coupon Schedules</h1>
  <div class="ts">Upload a file with Maturity Date, ISIN, Series, Face Value → all other columns calculated automatically</div>
</div>
""", unsafe_allow_html=True)

tab_xl, tab_csv = st.tabs(["📊  Excel Upload", "📋  CSV Upload"])

def handle_df(df):
    results = []
    for _, row in df.iterrows():
        try:
            r = process_row(row.to_dict())
            results.append(r)
        except Exception as e:
            st.warning(f"Skipped a row due to error: {e}")
    return results

with tab_xl:
    xl_file = st.file_uploader("Upload Excel file", type=["xlsx","xls"],
                                label_visibility="collapsed", key="xl_up")
    st.caption("Required columns: **Maturity Date**, **ISIN**, **Series**, **Face Value (Rs Mn)** · Extra columns/headers are ignored automatically")
    if xl_file:
        df, status = parse_uploaded_file(xl_file)
        if status == "ok" and df is not None:
            results = handle_df(df)
            if results:
                render_dashboard(results, key_suffix="xl")
            else:
                st.markdown('<div class="warn-box">⚠ No valid bond rows found. Check column names.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="warn-box">⚠ Could not parse file: {status}</div>', unsafe_allow_html=True)

with tab_csv:
    csv_file = st.file_uploader("Upload CSV file", type=["csv"],
                                 label_visibility="collapsed", key="csv_up")
    st.caption("Required columns: **Maturity Date**, **ISIN**, **Series**, **Face Value (Rs Mn)** · Extra columns/headers are ignored automatically")
    if csv_file:
        df, status = parse_uploaded_file(csv_file)
        if status == "ok" and df is not None:
            results = handle_df(df)
            if results:
                render_dashboard(results, key_suffix="csv")
            else:
                st.markdown('<div class="warn-box">⚠ No valid bond rows found. Check column names.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="warn-box">⚠ Could not parse file: {status}</div>', unsafe_allow_html=True)